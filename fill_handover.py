#!/usr/bin/env python3
"""
fill_handover.py — fill the Shift Handover Excel TEMPLATE from extracted data,
without destroying the template's formatting.

How it works
------------
1. Load the existing template with openpyxl (formatting, merged cells, styles
   all stay intact — we only overwrite specific cells).
2. For each field, FIND the cell whose text equals the configured `anchor`
   label, then write the value at a fixed (row, col) OFFSET from that label.
   -> No hardcoded cell addresses, so it survives small layout changes.
3. If the target cell sits inside a merged range, resolve to the range's
   top-left anchor (openpyxl ignores writes to any other merged cell).
4. Save to a NEW file (never overwrites the blank template).

Data comes from a JSON file: { "<field key>": <value>, ... }
Produce that JSON either by hand, or with the LLM prompt in extract_prompt.txt.

Usage
-----
    # preview what it would write (no file saved):
    python3 fill_handover.py --template "Shift Handover Template.xlsx" --data data.json

    # actually write the filled copy:
    python3 fill_handover.py --template "Shift Handover Template.xlsx" --data data.json --write
"""
from __future__ import annotations

import argparse
import json
from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ----------------------------------------------------------------------------
# FIELD MAP — edit this to match your template.
#
# Each entry maps a JSON `key` (from data.json) to:
#   anchor : EXACT text of the label cell in the template
#            (copy it verbatim from `inspect_template.py` output)
#   dr, dc : row / column offset from the label to the cell to fill.
#            dr=1,dc=0  -> value sits one row BELOW the label  (header-on-top)
#            dr=0,dc=1  -> value sits one column RIGHT of label (label-on-left)
#
# Anchors are matched case-insensitively and whitespace-trimmed.
# ----------------------------------------------------------------------------
@dataclass(frozen=True)
class Field:
    key: str
    anchor: str
    dr: int
    dc: int


FIELD_MAP: list[Field] = [
    # --- top header block (label on left, value to the right) ---
    Field("unit",              "Unit",                     dr=0, dc=1),
    Field("project_name",      "Project Name",             dr=0, dc=1),
    Field("delivery_lead",     "Delivery Lead",            dr=0, dc=1),
    # Field("date",            "Date",                     dr=0, dc=1),
    # --- incident counts (header on top, number below) ---
    Field("incidents_raised",   "No of Incidents Raised",   dr=1, dc=0),
    Field("incidents_resolved", "No of Incidents Resolved", dr=1, dc=0),
    Field("incidents_open",     "No of Incidents Open",     dr=1, dc=0),
    #
    # ---------------------------------------------------------------------
    # Below are observed from the template screen but the EXACT label text /
    # offset must be confirmed: run `inspect_template.py`, copy the real
    # label strings here, then uncomment. dr=1,dc=0 = value below the header.
    # ---------------------------------------------------------------------
    # --- incident aging buckets ---
    # Field("incidents_aging_5_3", "No of Incidents (5-3)", dr=1, dc=0),
    # Field("incidents_aging_3_1", "No of Incidents (3-1)", dr=1, dc=0),
    # Field("incidents_aging_gt5", "No of Incidents (>5)",  dr=1, dc=0),
    # --- service request (SR) counts ---
    # Field("srs_raised",   "No of SRs Raised",   dr=1, dc=0),
    # Field("srs_resolved", "No of SRs Resolved", dr=1, dc=0),
    # Field("srs_open",     "No of SRs Open",     dr=1, dc=0),
    # --- SR aging buckets ---
    # Field("srs_aging_5_3", "No of SRs (5-3)", dr=1, dc=0),
    # Field("srs_aging_3_1", "No of SRs (3-1)", dr=1, dc=0),
    # Field("srs_aging_gt5", "No of SRs (>5)",  dr=1, dc=0),
    # --- comment sections (label on left, Comments/Details to the right) ---
    # Field("major_incident_updates", "Major Incident Updates", dr=0, dc=1),
    # Field("activities",              "Activities",             dr=0, dc=1),
    # Field("pending_tasks",           "Pending Tasks / Follow-ups", dr=0, dc=1),
]


def normalize(text: object) -> str:
    return str(text).strip().lower() if text is not None else ""


def build_merge_index(ws: Worksheet) -> dict[tuple[int, int], tuple[int, int]]:
    """Map every (row, col) inside a merged range -> that range's top-left anchor."""
    index: dict[tuple[int, int], tuple[int, int]] = {}
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                index[(r, c)] = (rng.min_row, rng.min_col)
    return index


def find_anchor(ws: Worksheet, anchor: str) -> tuple[int, int] | None:
    target = normalize(anchor)
    for row in ws.iter_rows():
        for cell in row:
            if normalize(cell.value) == target:
                return cell.row, cell.column
    return None


def fill_sheet(
    ws: Worksheet,
    data: dict[str, object],
    *,
    write: bool,
) -> tuple[int, list[str]]:
    merge_index = build_merge_index(ws)
    written = 0
    warnings: list[str] = []

    for field in FIELD_MAP:
        if field.key not in data:
            warnings.append(f"data.json missing key '{field.key}' — skipped")
            continue
        value = data[field.key]
        if value is None or value == "":
            warnings.append(f"'{field.key}' value is empty — skipped")
            continue

        loc = find_anchor(ws, field.anchor)
        if loc is None:
            warnings.append(f"anchor not found in sheet: {field.anchor!r} ('{field.key}')")
            continue

        r, c = loc
        tr, tc = r + field.dr, c + field.dc
        tr, tc = merge_index.get((tr, tc), (tr, tc))  # resolve merged anchor
        coord = f"{get_column_letter(tc)}{tr}"

        existing = ws.cell(row=tr, column=tc).value
        print(f"  {field.key:<22} -> {coord:<5} = {value!r}"
              + (f"   (was {existing!r})" if existing not in (None, "") else ""))

        if write:
            ws.cell(row=tr, column=tc, value=value)
        written += 1

    return written, warnings


def main() -> None:
    ap = argparse.ArgumentParser(description="Fill Shift Handover template from JSON.")
    ap.add_argument("--template", required=True, help="path to the blank .xlsx template")
    ap.add_argument("--data", required=True, help="path to data.json")
    ap.add_argument("--sheet", default=None, help="sheet name (default: active sheet)")
    ap.add_argument("--out", default=None, help="output path (default: <template>.filled.xlsx)")
    ap.add_argument("--write", action="store_true", help="actually write the file (default: dry run)")
    args = ap.parse_args()

    with open(args.data, encoding="utf-8") as fh:
        data = json.load(fh)

    wb = load_workbook(args.template, data_only=False)
    ws = wb[args.sheet] if args.sheet else wb.active

    print(f"Template : {args.template}")
    print(f"Sheet    : {ws.title!r}")
    print(f"Mode     : {'WRITE' if args.write else 'DRY RUN (no file saved)'}")
    print("-" * 60)

    written, warnings = fill_sheet(ws, data, write=args.write)

    print("-" * 60)
    print(f"{written} cell(s) {'written' if args.write else 'would be written'}.")
    for w in warnings:
        print(f"  ! {w}")

    if args.write:
        out = args.out or args.template.rsplit(".", 1)[0] + ".filled.xlsx"
        wb.save(out)
        print(f"\nSaved -> {out}")
    else:
        print("\n(dry run — re-run with --write to save the filled copy)")


if __name__ == "__main__":
    main()
