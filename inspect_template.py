#!/usr/bin/env python3
"""
inspect_template.py — dump the structure of an Excel template.

Run this FIRST. It prints every non-empty cell (coordinate + text) and every
merged range for each sheet. Use the output to copy the EXACT label texts into
FIELD_MAP in fill_handover.py.

Usage:
    python3 inspect_template.py "Shift Handover Template.xlsx"
"""
import sys
from openpyxl import load_workbook


def main(path: str) -> None:
    wb = load_workbook(path, data_only=False)
    for ws in wb.worksheets:
        print("=" * 70)
        print(f"SHEET: {ws.title!r}   dims={ws.dimensions}")
        print("=" * 70)

        merged = sorted(str(r) for r in ws.merged_cells.ranges)
        if merged:
            print("MERGED RANGES:", ", ".join(merged))
            print("-" * 70)

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value).strip()
                if text:
                    # quotes make trailing spaces / hidden chars visible
                    print(f"{cell.coordinate:>5}  {text!r}")
        print()


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit("usage: python3 inspect_template.py <template.xlsx>")
    main(sys.argv[1])
